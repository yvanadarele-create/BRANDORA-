#!/usr/bin/env python3
"""
Build the Brandora Union manufacturer-sourcing workbook.

The workbook is a container for information the founder gathers herself on
Made-in-China and elsewhere. It contains **no supplier data**: no company
names, no prices, no contacts, no ratings, no shipping costs. Every data cell
ships blank, because a plausible-looking placeholder in a sourcing sheet is
indistinguishable from a real quote three weeks later, and that is how people
end up wiring money to a factory that was never verified.

Two conventions run through the whole file and are worth stating once:

  * **Blank means unknown. Zero means confirmed zero.**
    A missing shipping cost is not free shipping. The totals refuse to
    compute rather than quietly understate a landed cost. If a supplier has
    actually confirmed there is no tooling charge, type 0 — that is a fact,
    and the formula will use it.

  * **Every sheet ends in "Evidence / What I Actually Saw".**
    Not a summary, not an impression: what was on the screen or in the
    message. It is the column that separates a quote from a memory of one.

Run:  python3 tools/build-sourcing-workbook.py
Out:  docs/brandora-sourcing-workbook.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# House style — the identity, not decoration for its own sake.
# ---------------------------------------------------------------------------

PURPLE = "35105C"          # --metallic-purple
PURPLE_SOFT = "7650A5"     # --soft-purple
COMPUTED_FILL = "F2EEF7"   # formula columns, so they are visibly not for typing
EVIDENCE_FILL = "FBF7EE"   # the evidence column, warm, so the eye finds it

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=PURPLE)
BODY_FONT = Font(name="Calibri", size=11)

THIN = Side(style="thin", color="D9D2E4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROWS = 200  # pre-formatted rows, so the sheet is usable from the first entry


def header_row(ws, columns):
    """Write the header, style it, freeze it, and turn on the filter."""
    for index, spec in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=index, value=spec["name"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
        if spec.get("note"):
            note = Comment(spec["note"], "Brandora Union")
            note.width = 320
            note.height = 130
            cell.comment = note
        ws.column_dimensions[get_column_letter(index)].width = spec.get("width", 20)

    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ROWS + 1}"


def body(ws, columns):
    """Apply body formatting and any per-column formula, row by row."""
    for row in range(2, ROWS + 2):
        for index, spec in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=index)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=spec.get("wrap", False))

            if spec.get("formula"):
                cell.value = spec["formula"].format(r=row)
                cell.fill = PatternFill("solid", fgColor=COMPUTED_FILL)
            elif spec["name"].startswith("Evidence"):
                cell.fill = PatternFill("solid", fgColor=EVIDENCE_FILL)

            if spec.get("format"):
                cell.number_format = spec["format"]

        ws.row_dimensions[row].height = 30


def dropdown(ws, column_index, options, columns):
    """
    Attach a list validation to one column for the whole sheet.

    Deliberately a *warning*, not a block. These lists are guesses at the
    shapes a supplier conversation takes, and the founder will meet ones I did
    not think of. A spreadsheet that refuses to record what actually happened
    trains people to record something else.
    """
    letter = get_column_letter(column_index)
    validation = DataValidation(
        type="list",
        formula1='"' + ",".join(options) + '"',
        allow_blank=True,
        showDropDown=False,  # openpyxl inverts this: False shows the arrow
    )
    validation.errorStyle = "warning"
    validation.showErrorMessage = True
    validation.errorTitle = "Not one of the options"
    validation.error = (
        "That is not one of the listed values. Keep it if it is what really "
        "happened — the list is a shortcut, not a rule."
    )
    validation.showInputMessage = True
    validation.promptTitle = "Options"
    validation.prompt = "\n".join(options)
    ws.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}{ROWS + 1}")


def score_validation(ws, first, last):
    """
    Scores are whole numbers 1-5.

    This one *does* block, because a 7 out of 5 is never a judgement — it is a
    slipped finger, and it would quietly inflate a total that decides who gets
    the order.
    """
    validation = DataValidation(
        type="whole", operator="between", formula1=1, formula2=5, allow_blank=True
    )
    validation.errorStyle = "stop"
    validation.showErrorMessage = True
    validation.errorTitle = "Score out of range"
    validation.error = (
        "Scores run from 1 to 5. Leave the cell blank if you have not assessed "
        "this yet — blank is honest, a guessed 3 is not."
    )
    validation.showInputMessage = True
    validation.promptTitle = "Score 1-5"
    validation.prompt = "1 = clearly weak, 3 = acceptable, 5 = clearly strong.\nBlank = not yet assessed."
    ws.add_data_validation(validation)
    for index in range(first, last + 1):
        letter = get_column_letter(index)
        validation.add(f"{letter}2:{letter}{ROWS + 1}")


def col(name, width=20, wrap=False, formula=None, fmt=None, note=None):
    return {"name": name, "width": width, "wrap": wrap, "formula": formula,
            "format": fmt, "note": note}


DATE = "yyyy-mm-dd"
MONEY = "#,##0.00"

EVIDENCE_NOTE = (
    "What you actually saw, in your own words.\n\n"
    "Examples: \"Profile page, 12 Aug 2026 — states 15 years on the platform, "
    "audited supplier badge visible\" or \"WhatsApp message from Mr Chen, "
    "10 Aug — quoted USD 0.42/unit at 5,000 units, did not mention shipping.\"\n\n"
    "If you inferred it rather than saw it, say so. This column is what "
    "separates a quote from a memory of one."
)

# ---------------------------------------------------------------------------
# 1. Manufacturer Database — 44 columns
# ---------------------------------------------------------------------------

MANUFACTURERS = [
    # Identity and where it came from
    col("Supplier ID", 12, note="Your own reference, e.g. SUP-001. Every other sheet points back to this."),
    col("Company Name (as listed)", 32, wrap=True,
        note="Copy the name exactly as the platform shows it, including any Co., Ltd. Do not tidy it up — you will need to match it later."),
    col("Made-in-China Profile URL", 34, wrap=True),
    col("Date Found", 13, fmt=DATE),
    col("Found Via (search term / category)", 26, wrap=True),
    col("Country", 14),
    col("Province / Region", 18),
    col("City", 16),
    col("Full Address (as listed)", 34, wrap=True),
    col("Business Type", 20,
        note="Manufacturer, Trading Company, or both. A trading company is not a factory; it can still be the right choice, but the price will carry a margin you cannot negotiate away at the source."),
    col("Year Established", 14),
    col("Number of Employees (as listed)", 20),
    col("Factory Size (as listed)", 20),
    col("Main Products (as listed)", 34, wrap=True),
    col("Product Categories Relevant to Brandora", 30, wrap=True),
    col("Annual Revenue (as listed)", 20),
    col("Main Export Markets (as listed)", 26, wrap=True),
    col("Export Experience to Africa", 22,
        note="Yes / No / Not stated. \"Not stated\" is the honest answer when the profile is silent — it is not the same as No."),
    col("Export Experience to Côte d'Ivoire", 24),

    # Verification and trust
    col("Audited Supplier", 18),
    col("Verification Badge / Level (as listed)", 26, wrap=True),
    col("Years on Made-in-China (as listed)", 22),
    col("Certifications Claimed", 30, wrap=True,
        note="What the supplier says it holds — ISO 9001, BRC, FDA, FSC, and so on. Claimed is not the same as verified; that is the next column."),
    col("Certificate Evidence Seen", 22,
        note="Did you see the actual certificate document, with a number and an expiry date? A logo on a profile page is not evidence."),
    col("Trade Assurance / Escrow Available", 26),
    col("Third-party Inspection Accepted", 26),

    # Capability
    col("MOQ (as listed)", 16),
    col("MOQ Unit", 14),
    col("Customization Offered", 26, wrap=True),
    col("Printing Methods Offered", 26, wrap=True),
    col("Sample Available", 18),
    col("Sample Cost (as listed)", 20),
    col("Sample Lead Time (days)", 20),
    col("Production Lead Time (days)", 22),
    col("Packaging Materials Handled", 28, wrap=True),
    col("Food-grade Capability", 20),

    # Commercial and contact
    col("Payment Terms (as listed)", 26, wrap=True),
    col("Accepted Payment Methods", 26, wrap=True),
    col("Incoterms Offered", 20,
        note="EXW, FOB, CIF, DDP. The Incoterm decides who pays for what and where the risk transfers — a cheaper EXW price can land dearer than a higher CIF one."),
    col("Nearest Port", 18),
    col("Contact Person", 20),
    col("Contact Email", 26),
    col("Contact Phone / WhatsApp", 24),
    col("Evidence / What I Actually Saw", 46, wrap=True, note=EVIDENCE_NOTE),
]

# ---------------------------------------------------------------------------
# 2. Product / Quote Tracker — 25 columns, two computed
# ---------------------------------------------------------------------------

PRODUCT_TOTAL_NOTE = (
    "Computed: Quantity × Unit Price.\n\n"
    "Stays blank until both are filled in. Do not type over it — enter the "
    "quantity and the unit price and the total follows."
)

TOTAL_COST_NOTE = (
    "Computed: Product Total + Shipping + Customization / Tooling.\n\n"
    "Stays blank unless all three are present. A missing shipping cost is NOT "
    "free shipping — the total refuses to compute rather than understate what "
    "the goods will actually cost you landed.\n\n"
    "If a supplier has confirmed there is genuinely no tooling or shipping "
    "charge, type 0. Zero is a fact you were told. Blank is a question you "
    "have not asked yet."
)

QUOTES = [
    col("Quote ID", 12, note="Your own reference, e.g. Q-001."),
    col("Supplier ID", 12, note="Must match a Supplier ID on the Manufacturer Database sheet."),
    col("Supplier Name", 30, wrap=True),
    col("Date of Quote", 14, fmt=DATE),
    col("Product Name", 28, wrap=True),
    col("Product Category", 22),
    col("Material", 22, wrap=True),
    col("Size / Dimensions", 22, wrap=True),
    col("Capacity / Volume", 18),
    col("Colour", 16),
    col("Print / Customization Requested", 30, wrap=True),
    col("MOQ for This Product", 20),
    col("Quantity Quoted", 16),
    col("Unit Price", 14, fmt=MONEY),
    col("Currency", 12, note="USD, CNY, EUR, XOF. Record the currency the supplier quoted in, not your conversion of it."),
    col("Product Total", 16, fmt=MONEY, note=PRODUCT_TOTAL_NOTE,
        formula='=IF(OR(M{r}="",N{r}=""),"",M{r}*N{r})'),
    col("Shipping Cost", 16, fmt=MONEY,
        note="Leave blank until you have an actual figure. Blank is honest; zero is a claim."),
    col("Shipping Method", 18),
    col("Customization / Tooling Cost", 24, fmt=MONEY,
        note="Plates, moulds, dies, setup. Often quoted separately and forgotten — it is what makes a first order dearer per unit than a repeat one."),
    col("Total Estimated Cost", 20, fmt=MONEY, note=TOTAL_COST_NOTE,
        formula='=IF(OR(P{r}="",Q{r}="",S{r}=""),"",P{r}+Q{r}+S{r})'),
    col("Sample Cost", 14, fmt=MONEY),
    col("Lead Time (days)", 16),
    col("Incoterm", 14),
    col("Quote Valid Until", 18, fmt=DATE,
        note="Most factory quotes expire in 15–30 days. A quote you cannot date is a quote you cannot rely on."),
    col("Evidence / What I Actually Saw", 46, wrap=True, note=EVIDENCE_NOTE),
]

# ---------------------------------------------------------------------------
# 3. Contact / Follow-up Tracker — 14 columns, one dropdown
# ---------------------------------------------------------------------------

STATUSES = [
    "Not contacted",
    "Contacted - awaiting reply",
    "Replied - reviewing",
    "Quote requested",
    "Quote received",
    "Sample requested",
    "Sample received",
    "Negotiating",
    "Approved supplier",
    "On hold",
    "Rejected",
]

CONTACTS = [
    col("Contact ID", 12),
    col("Supplier ID", 12, note="Must match a Supplier ID on the Manufacturer Database sheet."),
    col("Supplier Name", 30, wrap=True),
    col("Contact Person", 20),
    col("Contact Method", 20),
    col("Date Contacted", 15, fmt=DATE),
    col("Message Sent (summary)", 40, wrap=True),
    col("Response Received", 18),
    col("Date of Response", 16, fmt=DATE),
    col("Response Summary", 40, wrap=True),
    col("Status", 24, note="Pick from the list. The list is deliberately short — a status nobody can define is a status nobody updates."),
    col("Next Action", 34, wrap=True),
    col("Next Follow-up Date", 19, fmt=DATE),
    col("Evidence / What I Actually Saw", 46, wrap=True, note=EVIDENCE_NOTE),
]

# ---------------------------------------------------------------------------
# 4. Supplier Comparison — seven dimensions out of five, thirty-five total
# ---------------------------------------------------------------------------

SCORE_NOTE = (
    "Score 1–5, or leave blank if you have not assessed it yet.\n\n"
    "1 = clearly weak  ·  3 = acceptable  ·  5 = clearly strong\n\n"
    "Score what you have evidence for. An unassessed dimension is blank, "
    "not a 3."
)

TOTAL_SCORE_NOTE = (
    "Computed: the sum of all seven dimensions, out of 35.\n\n"
    "Stays blank until all seven are scored. A total built from four scores "
    "reads like a low total and ranks the supplier below one you simply "
    "looked at for longer — so it does not compute at all until the "
    "comparison is fair."
)

COMPARISON = [
    col("Supplier ID", 12),
    col("Supplier Name", 30, wrap=True),
    col("Product / Quote Ref", 20, note="Which quote this comparison is about. Comparing suppliers without fixing the product compares nothing."),
    col("Price Competitiveness (/5)", 22, note=SCORE_NOTE),
    col("Product Quality Evidence (/5)", 24, note=SCORE_NOTE),
    col("Customization Capability (/5)", 24, note=SCORE_NOTE),
    col("Communication & Responsiveness (/5)", 28, note=SCORE_NOTE),
    col("Certifications & Compliance (/5)", 26, note=SCORE_NOTE),
    col("Lead Time & Reliability (/5)", 24, note=SCORE_NOTE),
    col("Shipping & Logistics to Côte d'Ivoire (/5)", 30, note=SCORE_NOTE),
    col("Total Score (/35)", 16, note=TOTAL_SCORE_NOTE,
        formula='=IF(COUNT(D{r}:J{r})=7,SUM(D{r}:J{r}),"")'),
    col("Scoring Basis / Evidence", 46, wrap=True,
        note="Why these scores. A number without a reason is not a comparison, it is a mood."),
    col("Decision", 22),
    col("Decision Date", 15, fmt=DATE),
]

DECISIONS = [
    "Shortlisted",
    "Sample ordered",
    "Approved",
    "Held for later",
    "Rejected",
    "Undecided",
]

# ---------------------------------------------------------------------------
# 5. Brandora Sourcing Notes — 10 columns
# ---------------------------------------------------------------------------

TOPICS = [
    "Packaging",
    "Pricing",
    "Logistics",
    "Compliance",
    "Branding",
    "Quality",
    "Risk",
    "Other",
]

NOTE_STATUSES = ["Open", "In progress", "Closed"]

NOTES = [
    col("Note ID", 12),
    col("Date", 13, fmt=DATE),
    col("Related Supplier ID", 18),
    col("Related Quote ID", 18),
    col("Topic", 18),
    col("Observation", 46, wrap=True, note="What you noticed. Facts first — the interpretation goes in the next column."),
    col("Why It Matters for Brandora", 42, wrap=True),
    col("Source / Where I Saw It", 34, wrap=True),
    col("Action Required", 34, wrap=True),
    col("Status", 16),
]


def build(path: Path) -> None:
    book = Workbook()
    book.remove(book.active)

    sheets = [
        ("Manufacturer Database", MANUFACTURERS),
        ("Product & Quote Tracker", QUOTES),
        ("Contact & Follow-up", CONTACTS),
        ("Supplier Comparison", COMPARISON),
        ("Brandora Sourcing Notes", NOTES),
    ]

    for title, columns in sheets:
        sheet = book.create_sheet(title)
        sheet.sheet_properties.tabColor = PURPLE_SOFT
        header_row(sheet, columns)
        body(sheet, columns)

    manufacturers = book["Manufacturer Database"]
    dropdown(manufacturers, 10, ["Manufacturer", "Trading Company", "Manufacturer & Trading", "Not stated"], MANUFACTURERS)
    for index in (18, 19, 20, 24, 25, 26, 31, 36):
        dropdown(manufacturers, index, ["Yes", "No", "Not stated"], MANUFACTURERS)

    quotes = book["Product & Quote Tracker"]
    dropdown(quotes, 18, ["Sea freight", "Air freight", "Express courier", "Not stated"], QUOTES)

    contacts = book["Contact & Follow-up"]
    dropdown(contacts, 5, ["Email", "WhatsApp", "Platform message", "Phone call", "Video call"], CONTACTS)
    dropdown(contacts, 8, ["Yes", "No"], CONTACTS)
    dropdown(contacts, 11, STATUSES, CONTACTS)

    comparison = book["Supplier Comparison"]
    score_validation(comparison, 4, 10)
    dropdown(comparison, 13, DECISIONS, COMPARISON)

    notes = book["Brandora Sourcing Notes"]
    dropdown(notes, 5, TOPICS, NOTES)
    dropdown(notes, 10, NOTE_STATUSES, NOTES)

    book.properties.title = "Brandora Union — Manufacturer Sourcing Workbook"
    book.properties.creator = "Brandora Union"
    book.properties.description = (
        "Empty by design. Blank means unknown; zero means confirmed zero. "
        "Totals refuse to compute on missing inputs rather than understate a cost."
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)


if __name__ == "__main__":
    out = Path(__file__).resolve().parents[1] / "docs" / "brandora-sourcing-workbook.xlsx"
    build(out)
    print(f"wrote {out}")
