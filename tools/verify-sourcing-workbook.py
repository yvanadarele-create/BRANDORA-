#!/usr/bin/env python3
"""
Check the sourcing workbook against what was asked for.

Two things matter here and nothing else does. First, the shape: five sheets,
the right number of columns on each, an evidence column where one was
promised, a working status dropdown. Second — and this is the point of the
whole file — that the workbook contains no invented supplier information. A
single stray example company name would undo it.

The formula behaviour is checked by recalculating the real file in
LibreOffice, not by reading the formula strings back and trusting them. The
cases that matter are the ones where something is missing: a quantity with no
price, a product total with no shipping. Those must come back empty, because
a total that treats a missing shipping cost as zero is worse than no total.
"""

import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / "docs" / "brandora-sourcing-workbook.xlsx"

EXPECTED = {
    "Manufacturer Database": 44,
    "Product & Quote Tracker": 25,
    "Contact & Follow-up": 14,
    "Supplier Comparison": 14,
    "Brandora Sourcing Notes": 10,
}

failures = []
checks = []


def check(name, ok, detail=""):
    checks.append((name, ok, detail))
    if not ok:
        failures.append(f"{name} — {detail}")


book = load_workbook(BOOK)

check("five sheets", book.sheetnames == list(EXPECTED), str(book.sheetnames))

for title, expected_columns in EXPECTED.items():
    if title not in book.sheetnames:
        continue
    sheet = book[title]
    headers = [c.value for c in sheet[1]]
    filled = [h for h in headers if h]
    check(f"{title}: {expected_columns} columns", len(filled) == expected_columns,
          f"found {len(filled)}")
    check(f"{title}: no blank header", all(headers[: len(filled)]), "")
    check(f"{title}: header frozen", sheet.freeze_panes == "A2", str(sheet.freeze_panes))

# The whole point: no supplier data anywhere.
formula_cells = 0
stray = []
for title in book.sheetnames:
    sheet = book[title]
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None:
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells += 1
                continue
            stray.append(f"{title}!{cell.coordinate} = {cell.value!r}")

check("no invented data in any cell", not stray, "; ".join(stray[:5]))
check("formulas are present", formula_cells > 0, f"{formula_cells} formula cells")

# The evidence column, where it was promised.
for title in ["Manufacturer Database", "Product & Quote Tracker", "Contact & Follow-up"]:
    headers = [c.value for c in book[title][1]]
    check(f"{title}: evidence column", "Evidence / What I Actually Saw" in headers, "")

# The status dropdown.
contacts = book["Contact & Follow-up"]
status_lists = [
    dv for dv in contacts.data_validations.dataValidation
    if dv.type == "list" and "Quote received" in (dv.formula1 or "")
]
check("contact status is a dropdown", len(status_lists) == 1,
      f"{len(status_lists)} matching validations")
if status_lists:
    status = status_lists[0]
    check("status dropdown covers column K", str(status.sqref).startswith("K2:"), str(status.sqref))
    check("status dropdown arrow is visible", status.showDropDown is False, str(status.showDropDown))
    check("status dropdown warns rather than blocks", status.errorStyle == "warning"
          and status.showErrorMessage, f"{status.errorStyle}, shown={status.showErrorMessage}")

# Scores constrained to 1-5.
comparison = book["Supplier Comparison"]
whole = [dv for dv in comparison.data_validations.dataValidation if dv.type == "whole"]
check("scores constrained 1-5", whole and str(whole[0].formula1) == "1"
      and str(whole[0].formula2) == "5",
      str([(dv.formula1, dv.formula2) for dv in whole]))
check("an out-of-range score is blocked, not merely flagged",
      whole and whole[0].errorStyle == "stop" and whole[0].showErrorMessage,
      str([(dv.errorStyle, dv.showErrorMessage) for dv in whole]))
check("scores cover all seven dimensions (D-J)",
      whole and str(whole[0].sqref).split() == [f"{c}2:{c}201" for c in "DEFGHIJ"],
      str(whole[0].sqref) if whole else "none")

# --- Recalculate for real -------------------------------------------------
#
# The formulas below are lifted out of the shipped workbook rather than
# retyped here, and evaluated by a spreadsheet formula engine rather than by
# anything of my own. Retyping them would only prove that I can copy a string;
# reimplementing IF and COUNT would only prove my reimplementation agrees with
# itself. What is under test is the formula the founder will actually open.

import formulas  # noqa: E402
from openpyxl import Workbook  # noqa: E402

quotes_source = book["Product & Quote Tracker"]
compare_source = book["Supplier Comparison"]

PRODUCT_TOTAL = quotes_source["P2"].value
LANDED_TOTAL = quotes_source["T2"].value
TOTAL_SCORE = compare_source["K2"].value

check("product total formula is the one specified",
      PRODUCT_TOTAL == '=IF(OR(M2="",N2=""),"",M2*N2)', PRODUCT_TOTAL)
check("landed total formula is the one specified",
      LANDED_TOTAL == '=IF(OR(P2="",Q2="",S2=""),"",P2+Q2+S2)', LANDED_TOTAL)
check("total score formula is the one specified",
      TOTAL_SCORE == '=IF(COUNT(D2:J2)=7,SUM(D2:J2),"")', TOTAL_SCORE)


def shift(formula: str, row: int) -> str:
    """Move a row-2 formula down to `row`, the way filling down does."""
    import re
    return re.sub(r"(?<=[A-Z])2\b", str(row), formula)


work = Path(tempfile.mkdtemp())
probe = work / "probe.xlsx"

live = Workbook()
q = live.active
q.title = "Product & Quote Tracker"
c = live.create_sheet("Supplier Comparison")

# Row 2: everything present. The totals must compute.
q["M2"], q["N2"], q["Q2"], q["S2"] = 5000, 0.42, 380, 120
# Row 3: a quantity but no unit price. The product total must stay empty.
q["M3"] = 5000
# Row 4: priced and tooling known, but shipping unknown. This is the case the
# founder asked for by name: a missing shipping cost is not free shipping.
q["M4"], q["N4"], q["S4"] = 1000, 1.5, 0
# Row 5: shipping confirmed as zero by the supplier. Zero is a fact, so it counts.
q["M5"], q["N5"], q["Q5"], q["S5"] = 200, 2, 0, 0

for row in range(2, 6):
    q[f"P{row}"] = shift(PRODUCT_TOTAL, row)
    q[f"T{row}"] = shift(LANDED_TOTAL, row)

# Row 2: all seven dimensions scored. Row 3: six of the seven.
for column, value in zip("DEFGHIJ", [4, 5, 3, 4, 2, 3, 5]):
    c[f"{column}2"] = value
for column, value in zip("DEFGHI", [4, 5, 3, 4, 2, 3]):
    c[f"{column}3"] = value
for row in (2, 3):
    c[f"K{row}"] = shift(TOTAL_SCORE, row)

live.save(probe)

solution = formulas.ExcelModel().loads(str(probe)).finish().calculate()
book_key = probe.name


def value_at(sheet, coordinate):
    key = f"'[{book_key}]{sheet.upper()}'!{coordinate}"
    raw = solution[key].value[0, 0]
    return raw


EMPTY = ("", None)

check("product total computes when quantity and price are present",
      value_at("Product & Quote Tracker", "P2") == 2100,
      repr(value_at("Product & Quote Tracker", "P2")))
check("product total stays empty without a unit price",
      value_at("Product & Quote Tracker", "P3") in EMPTY,
      repr(value_at("Product & Quote Tracker", "P3")))
check("landed total computes when every input is present",
      value_at("Product & Quote Tracker", "T2") == 2600,
      repr(value_at("Product & Quote Tracker", "T2")))
check("landed total REFUSES to compute when shipping is unknown",
      value_at("Product & Quote Tracker", "T4") in EMPTY,
      repr(value_at("Product & Quote Tracker", "T4")))
check("landed total does compute when shipping is a confirmed zero",
      value_at("Product & Quote Tracker", "T5") == 400,
      repr(value_at("Product & Quote Tracker", "T5")))
check("total score computes when all seven are scored",
      value_at("Supplier Comparison", "K2") == 26,
      repr(value_at("Supplier Comparison", "K2")))
check("total score stays empty on a partial comparison",
      value_at("Supplier Comparison", "K3") in EMPTY,
      repr(value_at("Supplier Comparison", "K3")))

shutil.rmtree(work, ignore_errors=True)

width = max(len(n) for n, _, _ in checks)
for name, ok, detail in checks:
    mark = "ok  " if ok else "FAIL"
    print(f"{mark} {name.ljust(width)}  {detail}")

print()
print(f"{len(checks) - len(failures)}/{len(checks)} checks passed")
sys.exit(1 if failures else 0)
